#!/usr/bin/env python3
"""
DCF Model Validation Script
Validates Excel DCF models for formula errors and common DCF mistakes
"""

import sys
import json
from pathlib import Path
from typing import Optional


class DCFModelValidator:
    """Validates DCF models for errors and quality issues"""

    def __init__(self, excel_path: str):
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        self.excel_path = excel_path
        self.openpyxl = openpyxl

        if not Path(excel_path).exists():
            raise FileNotFoundError(f"File not found: {excel_path}")

        self.workbook_formulas = openpyxl.load_workbook(excel_path, data_only=False)
        self.workbook_values = openpyxl.load_workbook(excel_path, data_only=True)
        self.errors = []
        self.warnings = []
        self.info = []
        
    def validate_all(self) -> dict:
        """
        Run all validation checks

        Returns:
            Dict with validation results
        """
        from datetime import datetime

        self.check_sheet_structure()
        self.check_formula_errors()
        self.check_dcf_logic()

        results = {
            'file': self.excel_path,
            'validation_date': datetime.now().isoformat(),
            'status': 'PASS' if len(self.errors) == 0 else 'FAIL',
            'error_count': len(self.errors),
            'warning_count': len(self.warnings),
            'errors': self.errors,
            'warnings': self.warnings,
            'info': self.info
        }

        return results
    
    def check_sheet_structure(self):
        """Verify required sheets exist"""
        required_sheets = ['DCF', 'WACC', 'Sensitivity']
        sheet_names = self.workbook_values.sheetnames

        for sheet in required_sheets:
            if sheet not in sheet_names:
                self.warnings.append(f"Recommended sheet missing: {sheet}")
            else:
                self.info.append(f"Found sheet: {sheet}")

    def check_formula_errors(self):
        """Check for Excel formula errors in all sheets"""
        excel_errors = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']
        error_details = {err: [] for err in excel_errors}
        total_errors = 0
        total_formulas = 0

        for sheet_name in self.workbook_values.sheetnames:
            ws_values = self.workbook_values[sheet_name]
            ws_formulas = self.workbook_formulas[sheet_name]

            for row in ws_values.iter_rows():
                for cell in row:
                    formula_cell = ws_formulas[cell.coordinate]

                    # Count formulas
                    if formula_cell.value and isinstance(formula_cell.value, str) and formula_cell.value.startswith('='):
                        total_formulas += 1

                    # Check for errors
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                self.errors.append(f"{err} at {location}")
                                break

        # Add summary info
        self.info.append(f"Total formulas: {total_formulas}")
        if total_errors == 0:
            self.info.append("✓ No formula errors found")
        else:
            self.errors.append(f"Total formula errors: {total_errors}")

        return error_details, total_errors
    
    def check_dcf_logic(self):
        """Validate DCF-specific logic and calculations"""
        self._check_terminal_growth_vs_wacc()
        self._check_wacc_range()
        self._check_terminal_value_proportion()

    def _check_terminal_growth_vs_wacc(self):
        """Critical check: Terminal growth must be less than WACC"""
        try:
            dcf_sheet = self.workbook_values['DCF']

            terminal_growth = None
            wacc = None

            # Search for terminal growth and WACC values
            for row in dcf_sheet.iter_rows(max_row=100, max_col=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_str = cell.value.lower()
                        if 'terminal' in cell_str and 'growth' in cell_str:
                            # Look for value in adjacent cells
                            for offset in range(1, 5):
                                adjacent = dcf_sheet.cell(cell.row, cell.column + offset).value
                                if isinstance(adjacent, (int, float)) and 0 < adjacent < 1:
                                    terminal_growth = adjacent
                                    break
                        if 'wacc' in cell_str and wacc is None:
                            for offset in range(1, 5):
                                adjacent = dcf_sheet.cell(cell.row, cell.column + offset).value
                                if isinstance(adjacent, (int, float)) and 0 < adjacent < 1:
                                    wacc = adjacent
                                    break

            if terminal_growth is not None and wacc is not None:
                if terminal_growth >= wacc:
                    self.errors.append(
                        f"CRITICAL: Terminal growth ({terminal_growth:.2%}) >= WACC ({wacc:.2%}). "
                        "This creates infinite value and is mathematically invalid."
                    )
                else:
                    self.info.append(
                        f"✓ Terminal growth ({terminal_growth:.2%}) < WACC ({wacc:.2%})"
                    )
            else:
                self.warnings.append("Could not locate terminal growth and WACC values")

        except KeyError:
            self.warnings.append("DCF sheet not found")
        except Exception as e:
            self.warnings.append(f"Could not validate terminal growth vs WACC: {str(e)}")

    def _check_wacc_range(self):
        """Check if WACC is in reasonable range"""
        try:
            wacc_sheet = self.workbook_values.get('WACC') or self.workbook_values['DCF']
            wacc = None

            for row in wacc_sheet.iter_rows(max_row=100, max_col=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if 'wacc' in cell.value.lower():
                            for offset in range(1, 5):
                                adjacent = wacc_sheet.cell(cell.row, cell.column + offset).value
                                if isinstance(adjacent, (int, float)) and 0 < adjacent < 1:
                                    wacc = adjacent
                                    break

            if wacc is not None:
                if wacc < 0.05 or wacc > 0.20:
                    self.warnings.append(
                        f"WACC ({wacc:.2%}) is outside typical range (5%-20%). Verify calculation."
                    )
                else:
                    self.info.append(f"✓ WACC ({wacc:.2%}) in reasonable range")
            else:
                self.warnings.append("Could not locate WACC value")

        except Exception as e:
            self.warnings.append(f"Could not validate WACC range: {str(e)}")

    def _check_terminal_value_proportion(self):
        """Check if terminal value is reasonable proportion of enterprise value"""
        try:
            dcf_sheet = self.workbook_values['DCF']

            terminal_value = None
            enterprise_value = None

            for row in dcf_sheet.iter_rows(max_row=200, max_col=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_str = cell.value.lower()
                        if 'terminal' in cell_str and 'value' in cell_str and 'pv' in cell_str:
                            for offset in range(1, 5):
                                adjacent = dcf_sheet.cell(cell.row, cell.column + offset).value
                                if isinstance(adjacent, (int, float)) and adjacent > 0:
                                    terminal_value = adjacent
                                    break
                        if 'enterprise' in cell_str and 'value' in cell_str:
                            for offset in range(1, 5):
                                adjacent = dcf_sheet.cell(cell.row, cell.column + offset).value
                                if isinstance(adjacent, (int, float)) and adjacent > 0:
                                    enterprise_value = adjacent
                                    break

            if terminal_value is not None and enterprise_value is not None and enterprise_value > 0:
                proportion = terminal_value / enterprise_value
                if proportion > 0.80:
                    self.warnings.append(
                        f"Terminal value is {proportion:.1%} of EV (typically should be 50-70%). "
                        "Model may be over-reliant on terminal assumptions."
                    )
                elif proportion < 0.40:
                    self.warnings.append(
                        f"Terminal value is {proportion:.1%} of EV (typically should be 50-70%). "
                        "Check if terminal assumptions are too conservative."
                    )
                else:
                    self.info.append(f"✓ Terminal value is {proportion:.1%} of EV")
            else:
                self.warnings.append("Could not locate terminal value and enterprise value")

        except Exception as e:
            self.warnings.append(f"Could not validate terminal value proportion: {str(e)}")
    


def validate_dcf_model(excel_path: str) -> dict:
    """
    Validate a DCF model Excel file

    Args:
        excel_path: Path to Excel DCF model

    Returns:
        Dict with validation results
    """
    validator = DCFModelValidator(excel_path)
    return validator.validate_all()


def main():
    """Command-line interface"""
    if len(sys.argv) < 2:
        print("Usage: python validate_dcf.py <excel_file> [output.json]")
        print("\nValidates DCF model for:")
        print("  - Formula errors (#REF!, #DIV/0!, etc.)")
        print("  - Terminal growth < WACC (critical)")
        print("  - WACC in reasonable range (5-20%)")
        print("  - Terminal value proportion of EV (40-80%)")
        print("\nReturns JSON with errors, warnings, and info")
        print("\nExample: python validate_dcf.py model.xlsx")
        print("Example: python validate_dcf.py model.xlsx results.json")
        sys.exit(1)

    excel_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        results = validate_dcf_model(excel_file)

        # Print results
        print(json.dumps(results, indent=2))

        # Save to file if requested
        if output_file:
            with open(output_file, 'w') as f:
                json.dump(results, f, indent=2)

        # Exit with error code if validation failed
        sys.exit(0 if results['status'] == 'PASS' else 1)

    except Exception as e:
        error_result = {
            'file': excel_file,
            'status': 'ERROR',
            'error': str(e)
        }
        print(json.dumps(error_result, indent=2))
        sys.exit(1)


if __name__ == "__main__":
    main()
